#!/usr/bin/env python3
"""
Oblique Media Plan Generator
Usage: python generate_media_plan.py <config.json> <output.xlsx>
"""
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Colours ──────────────────────────────────────────────────────────────────
HEADER_BG   = "1C1C1E"  # Near-black header row
HEADER_FG   = "FFFFFF"  # White header text
TOTAL_BG    = "CCCCCC"  # Light grey total rows
BUDGET_BG   = "E8E8E8"  # Slightly lighter budget input rows
NOTE_FG     = "666666"  # Grey for disclaimer notes

STATUS_BG = {
    "Live":    "C6EFCE",  # Light green
    "Paused":  "E0E0E0",  # Light grey
    "KIV":     "FFEB9C",  # Light yellow
    "WIP OBL": "BDD7EE",  # Light blue
    "WIP":     "BDD7EE",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _font(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def _fill(color):
    return PatternFill("solid", fgColor=color) if color else None

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set(cell, value=None, bold=False, fg=None, bg=None, h="left", v="center",
         wrap=False, size=10, num_format=None):
    if value is not None:
        cell.value = value
    cell.font = _font(bold=bold, color=fg or "000000", size=size)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = _align(h=h, v=v, wrap=wrap)
    if num_format:
        cell.number_format = num_format

def _header_row(ws, row_num, labels: dict):
    """Apply header formatting to a dict of {coord: label}."""
    ws.row_dimensions[row_num].height = 30
    for coord, label in labels.items():
        _set(ws[coord], value=label, bold=True, fg=HEADER_FG, bg=HEADER_BG,
             h="center", v="center", wrap=True)

# ── Main ──────────────────────────────────────────────────────────────────────

def generate(config: dict, output_path: str):
    currency   = config.get("currency", "")
    cur_fmt    = f'"{currency} "#,##0'
    cur_fmt_dp = f'"{currency} "#,##0.00'

    wb = Workbook()
    ws = wb.active
    ws.title = "Media Plan"

    # Column widths: A=spacer, B=Channel, C=Status, D=CampaignType,
    #                E=CampaignName, F=Audience, G=Content, H=Bidding,
    #                I=MonthlyBudget, J=DailyBudget, K=%
    col_widths = {
        "A": 3.5, "B": 17, "C": 19, "D": 18, "E": 52,
        "F": 35,  "G": 27, "H": 22, "I": 15, "J": 11, "K": 9,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Title (row 3) ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 16
    title = f"Media Plan — {config.get('client_name', '')}"
    if config.get("cpa_target"):
        title += f"  |  Target CPA/CPL: {currency} {config['cpa_target']}"
    _set(ws["B3"], value=title, bold=True, h="left", size=12)

    # ── Budget Summary Row (row 4) ────────────────────────────────────────────
    ws.row_dimensions[4].height = 16
    total_planned = sum(ch.get("monthly_budget_input", 0) for ch in config.get("channels", []))
    channel_summary_parts = []
    for ch in config.get("channels", []):
        ch_budget = ch.get("monthly_budget_input", 0)
        pct = (ch_budget / total_planned * 100) if total_planned else 0
        ch_label = ch["name"].replace("Google Ads ", "Google ").replace("Meta Ads", "Meta")
        channel_summary_parts.append(f"{ch_label}: {currency} {ch_budget:,} ({pct:.0f}%)")
    summary_line = f"Total Monthly Budget: {currency} {total_planned:,}   |   " + "   |   ".join(channel_summary_parts)
    _set(ws["B4"], value=summary_line, bold=False, h="left", size=10, fg="444444")

    # ── Column Headers (row 6) ────────────────────────────────────────────────
    _header_row(ws, 6, {
        "B6": "Channel",
        "C6": "Status",
        "D6": "Campaign",
        "E6": "Campaign Name",
        "F6": "Audience",
        "G6": "Content",
        "H6": "Optimisation / Bidding Strategy",
        "I6": "Monthly Budget",
        "J6": "Daily Budget",
        "K6": "%",
    })

    # ── Channel Rows ──────────────────────────────────────────────────────────
    current_row   = 7
    channel_meta  = []   # (channel_name, total_row, budget_row)

    for channel in config.get("channels", []):
        ch_name    = channel["name"]
        ch_budget  = channel.get("monthly_budget_input", 0)
        campaigns  = channel.get("campaigns", [])
        ch_start   = current_row

        for camp in campaigns:
            r = current_row
            ws.row_dimensions[r].height = 19.5

            # Status (with colour)
            status = camp.get("status", "Paused")
            sc = ws[f"C{r}"]
            sc.value = status
            sc.font  = _font(size=10)
            sc.alignment = _align(h="center")
            if status in STATUS_BG:
                sc.fill = _fill(STATUS_BG[status])

            # Text columns
            for col, key in [("D", "campaign_type"), ("E", "campaign_name"),
                              ("F", "audience"), ("G", "content"), ("H", "bidding")]:
                c = ws[f"{col}{r}"]
                c.value     = camp.get(key, "")
                c.font      = _font(size=10)
                c.alignment = _align(h="left", wrap=True)

            # Budget columns
            monthly = camp.get("monthly_budget") or 0
            daily   = camp.get("daily_budget")
            if daily is None and monthly:
                daily = round(monthly / 31, 0)

            for col, val, fmt in [("I", monthly or "", cur_fmt),
                                   ("J", daily  or "", cur_fmt_dp)]:
                c = ws[f"{col}{r}"]
                c.value          = val
                c.font           = _font(size=10)
                c.alignment      = _align(h="center")
                c.number_format  = fmt

            current_row += 1

        ch_end = current_row - 1

        # Merge channel name column B across all campaign rows
        if ch_end >= ch_start:
            ws.merge_cells(f"B{ch_start}:B{ch_end}")
        bc = ws[f"B{ch_start}"]
        bc.value     = ch_name
        bc.font      = _font(size=10)
        bc.alignment = _align(h="left", v="center", wrap=True)

        # Total row
        tr = current_row
        ws.row_dimensions[tr].height = 19.5
        for col in "BCDEFGHIJK":
            ws[f"{col}{tr}"].fill      = _fill(TOTAL_BG)
            ws[f"{col}{tr}"].font      = _font(bold=True, size=10)
            ws[f"{col}{tr}"].alignment = _align(h="center")
        ws[f"H{tr}"].value = f"Total {ch_name}"
        ws[f"H{tr}"].alignment = _align(h="left")

        # J total = SUMIFS Live daily budgets
        ws[f"J{tr}"].value          = f'=SUMIFS(J{ch_start}:J{ch_end},C{ch_start}:C{ch_end},"Live")'
        ws[f"J{tr}"].number_format  = cur_fmt

        # I total = J * 30
        ws[f"I{tr}"].value          = f"=J{tr}*30"
        ws[f"I{tr}"].number_format  = cur_fmt

        current_row += 1

        # Monthly budget input row
        br = current_row
        ws.row_dimensions[br].height = 19.5
        for col in "BCDEFGHIJK":
            ws[f"{col}{br}"].fill      = _fill(BUDGET_BG)
            ws[f"{col}{br}"].font      = _font(bold=True, size=10)
            ws[f"{col}{br}"].alignment = _align(h="center")

        ch_label = ch_name.replace("Google Ads ", "Google ").replace("Meta Ads", "Meta")
        ws[f"H{br}"].value     = f"Monthly budget {ch_label.lower()}"
        ws[f"H{br}"].alignment = _align(h="left")
        ws[f"I{br}"].value          = ch_budget
        ws[f"I{br}"].number_format  = cur_fmt
        ws[f"J{br}"].value          = f"=I{br}/31"
        ws[f"J{br}"].number_format  = cur_fmt_dp

        # % formulas for each campaign in this channel
        for i, camp in enumerate(campaigns):
            r  = ch_start + i
            kc = ws[f"K{r}"]
            kc.value          = f'=IFERROR(J{r}/$J${br},"")'
            kc.font           = _font(size=10)
            kc.alignment      = _align(h="center")
            kc.number_format  = "0.0%"

        channel_meta.append((ch_name, tr, br))
        current_row += 2   # blank spacer row between channels

    # ── Grand Total ───────────────────────────────────────────────────────────
    # Two rows: Live Total (SUMIFS-based) and Planned Total (budget inputs)
    grand_row = current_row
    ws.row_dimensions[grand_row].height = 19.5
    for col in "BCDEFGHIJK":
        ws[f"{col}{grand_row}"].fill      = _fill(TOTAL_BG)
        ws[f"{col}{grand_row}"].font      = _font(bold=True, size=10)
        ws[f"{col}{grand_row}"].alignment = _align(h="center")

    ws[f"B{grand_row}"].value     = "Total (Live)"
    ws[f"B{grand_row}"].alignment = _align(h="left")

    i_parts = "+".join(f"I{tr}" for _, tr, _ in channel_meta)
    j_parts = "+".join(f"J{tr}" for _, tr, _ in channel_meta)
    ws[f"I{grand_row}"].value          = f"={i_parts}"
    ws[f"I{grand_row}"].number_format  = cur_fmt
    ws[f"J{grand_row}"].value          = f"={j_parts}"
    ws[f"J{grand_row}"].number_format  = cur_fmt

    # Planned total row — sums budget input rows (always shows full plan)
    planned_row = grand_row + 1
    ws.row_dimensions[planned_row].height = 19.5
    PLANNED_BG = "1C1C1E"
    for col in "BCDEFGHIJK":
        ws[f"{col}{planned_row}"].fill      = _fill(PLANNED_BG)
        ws[f"{col}{planned_row}"].font      = _font(bold=True, size=10, color="FFFFFF")
        ws[f"{col}{planned_row}"].alignment = _align(h="center")

    ws[f"B{planned_row}"].value     = "Total (Planned)"
    ws[f"B{planned_row}"].alignment = _align(h="left")

    bi_parts = "+".join(f"I{br}" for _, _, br in channel_meta)
    bj_parts = "+".join(f"J{br}" for _, _, br in channel_meta)
    ws[f"I{planned_row}"].value          = f"={bi_parts}"
    ws[f"I{planned_row}"].number_format  = cur_fmt
    ws[f"J{planned_row}"].value          = f"={bj_parts}"
    ws[f"J{planned_row}"].number_format  = cur_fmt_dp

    current_row += 1  # extra row for the planned total

    current_row += 2

    # ── Notes ─────────────────────────────────────────────────────────────────
    notes = config.get("notes", [
        "**Note: Cost per results may vary depending on audience size, competition & landing page experience.",
        "**Estimations are based on past campaigns performance & rough estimations.",
    ])
    for note in notes:
        nc = ws[f"B{current_row}"]
        nc.value     = note
        nc.font      = Font(size=9, italic=True, name="Arial", color=NOTE_FG)
        nc.alignment = _align(h="left")
        current_row += 1

    wb.save(output_path)
    print(json.dumps({"status": "success", "output": output_path}))


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python generate_media_plan.py <config.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1]) as f:
        cfg = json.load(f)
    generate(cfg, sys.argv[2])
