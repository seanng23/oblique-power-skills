"""
build_excel.py — Influencer Shortlist Excel Builder
Usage: python build_excel.py <output_path> <json_data_path>

Reads a JSON file of influencer records and writes a formatted Excel shortlist.

JSON input format (list of objects):
[
  {
    "handle": "username",
    "platform": "Instagram" | "TikTok",
    "profile_url": "https://...",
    "followers": 12500,
    "avg_engagement_rate": 3.4,   // as a percentage float, e.g. 3.4 means 3.4%
    "niche": "Lifestyle, Parenting",
    "rationale": "Active lifestyle creator based in KL..."
  },
  ...
]
"""

import sys
import json
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


def build_excel(output_path: str, data: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Influencer Shortlist"

    # ── Colours ────────────────────────────────────────────────────────────────
    HEADER_BG   = "1A1A2E"   # dark navy
    HEADER_FG   = "FFFFFF"   # white
    ALT_ROW_BG  = "F5F5F7"   # very light grey
    LINK_COLOR  = "0563C1"   # Excel hyperlink blue

    # ── Column definitions ─────────────────────────────────────────────────────
    COLUMNS = [
        ("#",               8),
        ("Handle",          22),
        ("Platform",        13),
        ("Profile URL",     38),
        ("Followers",       14),
        ("Avg. Engagement Rate", 22),
        ("Niche",           24),
        ("Rationale",       60),
    ]

    # ── Header row ─────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 28

    # ── Sort data: highest engagement first; N/A at bottom ────────────────────
    def sort_key(record):
        er = record.get("avg_engagement_rate")
        if er is None or er == "N/A":
            return -1
        try:
            return float(er)
        except (ValueError, TypeError):
            return -1

    data_sorted = sorted(data, key=sort_key, reverse=True)

    # ── Data rows ─────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_BG)
    link_font_base = Font(name="Calibri", size=10, color=LINK_COLOR, underline="single")
    normal_font    = Font(name="Calibri", size=10)

    thin_border = Border(
        bottom=Side(style="thin", color="E0E0E0")
    )

    for row_idx, record in enumerate(data_sorted, start=2):
        is_alt = (row_idx % 2 == 0)
        row_fill = alt_fill if is_alt else PatternFill(fill_type=None)

        # ── Engagement rate formatting ─────────────────────────────────────────
        er = record.get("avg_engagement_rate")
        if er is None or er == "N/A":
            er_display = "N/A"
        else:
            try:
                er_display = f"{float(er):.1f}%"
            except (ValueError, TypeError):
                er_display = "N/A"

        # ── Followers formatting ───────────────────────────────────────────────
        followers = record.get("followers", 0)
        try:
            followers_display = f"{int(followers):,}"
        except (ValueError, TypeError):
            followers_display = str(followers)

        row_values = [
            row_idx - 1,                            # #
            record.get("handle", ""),               # Handle
            record.get("platform", ""),             # Platform
            record.get("profile_url", ""),          # Profile URL
            followers_display,                      # Followers
            er_display,                             # Avg. Engagement Rate
            record.get("niche", ""),                # Niche
            record.get("rationale", ""),            # Rationale
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border

            if is_alt:
                cell.fill = alt_fill

            # Profile URL — make it a clickable hyperlink
            if col_idx == 4 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = link_font_base
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
            elif col_idx in (1, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = left_align

        # Taller rows for rationale
        ws.row_dimensions[row_idx].height = 45

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Summary stats row at the bottom ───────────────────────────────────────
    summary_row = len(data_sorted) + 3
    ws.cell(row=summary_row, column=1, value="Total influencers:").font = Font(bold=True, size=10)
    ws.cell(row=summary_row, column=2, value=len(data_sorted)).font    = Font(size=10)

    ig_count  = sum(1 for r in data_sorted if r.get("platform", "").lower() == "instagram")
    tt_count  = sum(1 for r in data_sorted if r.get("platform", "").lower() == "tiktok")
    ws.cell(row=summary_row+1, column=1, value="Instagram / TikTok:").font = Font(bold=True, size=10)
    ws.cell(row=summary_row+1, column=2, value=f"{ig_count} / {tt_count}").font = Font(size=10)

    er_values = []
    for r in data_sorted:
        er = r.get("avg_engagement_rate")
        if er and er != "N/A":
            try:
                er_values.append(float(er))
            except (ValueError, TypeError):
                pass

    if er_values:
        avg_er = sum(er_values) / len(er_values)
        ws.cell(row=summary_row+2, column=1, value="Avg. Engagement Rate:").font = Font(bold=True, size=10)
        ws.cell(row=summary_row+2, column=2, value=f"{avg_er:.1f}%").font = Font(size=10)

    wb.save(output_path)
    print(f"✅ Saved: {output_path}  ({len(data_sorted)} influencers)")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python build_excel.py <output_path.xlsx> <data.json>")
        sys.exit(1)

    output_path = sys.argv[1]
    json_path   = sys.argv[2]

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    build_excel(output_path, data)
